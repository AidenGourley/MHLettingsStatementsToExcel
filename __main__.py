import os
import time
import PyPDF2
import fiscalyear  # Change the fiscal year start point in the fiscalyear module
fiscalyear.START_DAY = 6
fiscalyear.START_MONTH = 4
from openpyxl import Workbook
from openpyxl import load_workbook
from termcolor import cprint
import shutil

import House
import Statement
import Transaction


def writeTableHeader(ws, fiscalYear):
    ws.title = str(fiscalYear)
    ws.cell(row=1, column=1).value = "FiscalYear: " + str(fiscalYear)
    ws.cell(row=2, column=1).value = "Month"
    ws.cell(row=2, column=2).value = "Gross Income"
    ws.cell(row=2, column=3).value = "Agent Fees"
    ws.cell(row=2, column=4).value = "Gardening Expenditure"
    ws.cell(row=2, column=5).value = "Other Expenditure"
    ws.cell(row=2, column=6).value = "Total Expenditure"
    ws.cell(row=2, column=7).value = "Retained"
    ws.cell(row=2, column=8).value = "Gross Profit"


def writeTransactionRow(ws, s, rowNum):
    ws.cell(row=rowNum, column=1).value = s.dateString
    ws.cell(row=rowNum, column=2).value = s.getGrossIncome()
    ws.cell(row=rowNum, column=3).value = s.agentFeesToString()
    ws.cell(row=rowNum, column=4).value = s.gardeningFeesToString()
    ws.cell(row=rowNum, column=5).value = s.otherFeesToString()
    ws.cell(row=rowNum, column=6).value = s.getGrossExpenditure()
    ws.cell(row=rowNum, column=7).value = s.retained
    ws.cell(row=rowNum, column=8).value = s.getGrossProfit()


def writeTotalRow(ws, year):
    # wb[Houses[house]].append()
    ws.cell(row=16, column=1).value = "TOTAL: "
    ws.cell(row=16, column=2).value = year.totalGrossIncome()
    ws.cell(row=16, column=3).value = year.totalAgentFees()
    ws.cell(row=16, column=4).value = year.totalGardeningFees()
    ws.cell(row=16, column=5).value = year.totalOtherExpenditure()
    ws.cell(row=16, column=6).value = year.totalGrossExpenditure()
    ws.cell(row=16, column=7).value = year.totalRetained()
    ws.cell(row=16, column=8).value = year.totalGrossProfit()


def assignStatementToHouse(Houses, statements):
    for statement in statements:
        for house in Houses:
            if house.getPostcode() in statement.address:
                house.addStatement(statement)
                break
    return Houses


def getFiscalMonth(statement, assignedRows):
    row = ((statement.startDate.month - 4) % 12) + 3
    if row in assignedRows:
        row += 1
    return row


def splitDate(startDate, endDate):
    endStatementStartDate = startDate
    endStatementEndDate = fiscalyear.FiscalDate(year=endDate.year, month=4, day=5)
    startStatementStartDate = fiscalyear.FiscalDate(year=endDate.year, month=4, day=6)
    startStatementEndDate = endDate
    return endStatementStartDate, endStatementEndDate, startStatementStartDate, startStatementEndDate


def splitStatement(statement):
    # Split the statements and assign new dates first.

    endStatement, startStatement = Statement.Statement(), Statement.Statement()
    endStatement.address, startStatement.address = statement.address, statement.address
    endStatement.startDate, endStatement.endDate, startStatement.startDate, startStatement.endDate = splitDate(
        statement.startDate, statement.endDate)
    differenceBetweenDates = abs((statement.endDate - statement.startDate).days)
    daysUntilNewFiscalYear = abs((endStatement.endDate - endStatement.startDate).days)
    endStatement.retained = round((statement.retained / differenceBetweenDates) * daysUntilNewFiscalYear, 2)
    startStatement.retained = round(
        (statement.retained / differenceBetweenDates) * (differenceBetweenDates - daysUntilNewFiscalYear), 2)
    startStatement.dateString = startStatement.startDate.strftime(
        "%d/%m/%Y") + " to " + startStatement.endDate.strftime("%d/%m/%Y")
    endStatement.dateString = endStatement.startDate.strftime("%d/%m/%Y") + " to " + endStatement.endDate.strftime(
        "%d/%m/%Y")
    print(statement.startDate, " - ", statement.endDate)
    print("---------------------------------------------")
    print(endStatement.startDate, " - ", endStatement.endDate)
    print(startStatement.startDate, " - ", startStatement.endDate)

    # Remove retained transaction from the endStatement, and keeps it in the startStatement
    for transaction in statement.incomeTransactions:
        if "refund" in transaction.name.lower():
            endStatement.incomeTransactions.append(transaction)
        else:
            # Split the transaction and assign it a weight. Add each new transaction back to the new statements.
            endNet = round((transaction.net / differenceBetweenDates) * daysUntilNewFiscalYear, 2)
            endGross = round((transaction.gross / differenceBetweenDates) * daysUntilNewFiscalYear, 2)
            endVAT = round((transaction.VAT / differenceBetweenDates) * daysUntilNewFiscalYear, 2)
            startNet = round(
                (transaction.net / differenceBetweenDates) * (differenceBetweenDates - daysUntilNewFiscalYear), 2)
            startGross = round((transaction.gross / differenceBetweenDates) * (
                    differenceBetweenDates - daysUntilNewFiscalYear), 2)
            startVAT = round(
                (transaction.VAT / differenceBetweenDates) * (differenceBetweenDates - daysUntilNewFiscalYear), 2)
            try:
                splitEndTransaction = Transaction.Transaction(transaction.name, endNet, endVAT, endGross, transaction.date)
                splitStartTransaction = Transaction.Transaction(transaction.name, startNet, startVAT, startGross,
                                                            transaction.date)
            except AttributeError:
                splitEndTransaction = Transaction.Transaction(transaction.name, endNet, endVAT, endGross,
                                                              None)
                splitStartTransaction = Transaction.Transaction(transaction.name, startNet, startVAT, startGross,
                                                                None)
            endStatement.incomeTransactions.append(splitEndTransaction)
            startStatement.incomeTransactions.append(splitStartTransaction)

    for transaction in statement.expenditureTransactions:
        endStatement.expenditureTransactions.append(transaction)

    endStatement.number, startStatement.number = statement.number

    return startStatement, endStatement


Houses = [House.House("RG45 6RY"), House.House("RG45 6JS")]

for filename in os.listdir("./testFiles/"):
    if filename.endswith(".pdf"):
        filePath = "./testFiles/" + filename
        pdfFileObj = open(filePath, 'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        stringFromPDF = ""
        for i in range(0, pdfReader.numPages):
            pageObj = pdfReader.getPage(i)
            stringFromPDF += pageObj.extractText()
        statement = Statement.Statement(stringFromPDF)
        #Check if split needed#
        startOfMonthFiscalYear = statement.startDate.fiscal_year
        endOfMonthFiscalYear = statement.endDate.fiscal_year
        if startOfMonthFiscalYear != endOfMonthFiscalYear:
            statement1, statement2 = splitStatement(statement)
            Houses = assignStatementToHouse(Houses, [statement1, statement2])
        else:
            Houses = assignStatementToHouse(Houses, [statement])


for house in Houses:
    wbName = house.getPostcode() + ".xlsx"
    years = house.getYears()
    try:
        wb = load_workbook(wbName)
    except:
        shutil.copy2('EmptyTemplate.xlsx', wbName)
        wb = Workbook(wbName)
        wb.close()
        wb = load_workbook(wbName)
    for year in years:
        print(len(year.getStatements()))
        try:
            ws = wb.get_sheet_by_name(str(year.getYear()))
        except:
            for i in wb.sheetnames:
                if "Sheet" in i:
                    ws = wb.get_sheet_by_name(i)
                    ws.title = str(year.getYear())
                    writeTableHeader(ws, year.getYear())
                    break
        writeTotalRow(ws, year)
        assignedRows = []
        for s in year.getStatements():
            row = getFiscalMonth(s, assignedRows)
            assignedRows.append(row)
            writeTransactionRow(ws, s, row)
            print(s.startDate, " - ", s.endDate)

    wb.save(wbName)


cprint("Spreadsheet Build Success!", "green")
input()
